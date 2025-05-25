import openpyxl
import pandas as pd
from textblob import TextBlob
import csv
from transformers import DistilBertTokenizer, DistilBertForSequenceClassification, Trainer, TrainingArguments
from datasets import Dataset
from sklearn.model_selection import train_test_split
import torch
import numpy as np

try:
    # sentiment analysis
    wb = openpyxl.Workbook()

    df = pd.read_csv('bbc-text.csv')
    
 
    sheet = wb.active
    sheet.title = "Sentiment Analysis"
    
  
    sheet['A1'] = 'Category'
    sheet['B1'] = 'Text'
    sheet['C1'] = 'Sentiment Score'
    

    for idx, row in df.iterrows():
        category = row['category']
        text = row['text']

        sentiment = TextBlob(text).sentiment.polarity
        

        sheet.cell(row=idx+2, column=1, value=category)
        sheet.cell(row=idx+2, column=2, value=text)
        sheet.cell(row=idx+2, column=3, value=sentiment)
    

    wb.save('sentiment_analysis.xlsx')
    print("Analysis complete and saved to sentiment_analysis.xlsx")
    
    #classification
    categories = df['category'].unique()
    print("\nUnique categories:", categories)

    category_keywords = {
        'tech': ['computer', 'software', 'internet', 'digital', 'technology', 'mobile', 'online'],
        'business': ['market', 'economy', 'company', 'stock', 'trade', 'financial', 'business'],
        'sport': ['game', 'team', 'player', 'match', 'tournament', 'championship', 'win'],
        'entertainment': ['film', 'movie', 'music', 'star', 'actor', 'show', 'entertainment'],
        'politics': ['government', 'election', 'party', 'minister', 'policy', 'vote', 'political']
    }

    def classify_text(text):
        scores = {}
        text = text.lower()
        for category, keywords in category_keywords.items():
            score = sum(text.count(keyword) for keyword in keywords)
            scores[category] = score
        return max(scores, key=scores.get)


    correct = 0
    predictions = []
    category_counts = {}
    total_articles = len(df)

    for idx, row in df.iterrows():
        predicted = classify_text(row['text'])
        actual = row['category']
        predictions.append(predicted)
        category_counts[actual] = category_counts.get(actual, 0) + 1
        if predicted == actual:
            correct += 1

    print("\nCategory Distribution:")
    for category, count in category_counts.items():
        percentage = (count / total_articles) * 100
        print(f"{category}: {count} articles ({percentage:.1f}%)")

    accuracy = (correct / total_articles) * 100
    print(f"\nClassification Accuracy: {accuracy:.1f}%")
    
    ## everything below this line is for the DistilBERT model and is AI.
    # Prepare data
    texts = df['text'].tolist()
    labels = df['category'].tolist()

    # Create label mapping
    label2id = {label: i for i, label in enumerate(set(labels))}
    id2label = {i: label for label, i in label2id.items()}
    num_labels = len(label2id)

    # Convert labels to ids
    label_ids = [label2id[label] for label in labels]

    # Split data
    train_texts, val_texts, train_labels, val_labels = train_test_split(
        texts, label_ids, test_size=0.2, random_state=42
    )

    # Initialize tokenizer and model
    tokenizer = DistilBertTokenizer.from_pretrained('distilbert-base-uncased')
    model = DistilBertForSequenceClassification.from_pretrained(
        'distilbert-base-uncased', 
        num_labels=num_labels
    )

    # Tokenize data
    def tokenize_function(examples):
        return tokenizer(examples, padding='max_length', truncation=True, max_length=512)

    train_encodings = tokenize_function(train_texts)
    val_encodings = tokenize_function(val_texts)

    # Create datasets
    train_dataset = Dataset.from_dict({
        'input_ids': train_encodings['input_ids'],
        'attention_mask': train_encodings['attention_mask'],
        'labels': train_labels
    })

    val_dataset = Dataset.from_dict({
        'input_ids': val_encodings['input_ids'],
        'attention_mask': val_encodings['attention_mask'],
        'labels': val_labels
    })

    # Training arguments
    training_args = TrainingArguments(
        output_dir="./news_classifier",
        num_train_epochs=3,
        per_device_train_batch_size=8,
        per_device_eval_batch_size=8,
        eval_steps=500,
        save_steps=500,
        logging_steps=100
    )

    # Initialize trainer
    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=train_dataset,
        eval_dataset=val_dataset,
    )

    # Train model
    trainer.train()

    # Evaluate
    results = trainer.evaluate()
    print(f"\nDistilBERT Model Evaluation Results:")
    print(f"Validation Loss: {results['eval_loss']:.3f}")

    # Make predictions on validation set
    predictions = trainer.predict(val_dataset)
    preds = np.argmax(predictions.predictions, axis=1)
    accuracy = (preds == val_labels).mean()
    print(f"DistilBERT Classification Accuracy: {accuracy * 100:.1f}%")
    ## ends here
except Exception as e:
    print(f"An error occurred: {e}")